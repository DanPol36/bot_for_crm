from aiogram import types
from aiogram.types import Message
from api_client import get_clients, get_person, get_orders, update_person, delete_person, add_client
import requests
from keyboards import main_keyboard
import io
import csv
from aiogram import types
from urllib.parse import unquote_plus, quote_plus
import os
import asyncio
import requests
import tempfile
from typing import Tuple

# In-memory simple state for interactive editing: user_id -> {'action':..., 'fio':...}
_USER_STATE: dict[int, dict] = {}
# In-memory auth: set of authorized telegram user ids
_AUTH_USERS: set[int] = set()

_ADMIN_LOGIN = os.getenv('ADMIN_LOGIN', 'admin')
_ADMIN_PASS = os.getenv('ADMIN_PASS', '12345')

def _is_authorized(user_id: int) -> bool:
    return user_id in _AUTH_USERS



async def start_handler(message: Message):
    if _is_authorized(message.from_user.id):
        await message.answer("Привет! Я CRM-бот — вы авторизованы.", reply_markup=main_keyboard)
    else:
        await message.answer("Привет! Я CRM-бот. Для управления выполните вход: используйте /login")


async def show_clients(message: Message):
    if not _is_authorized(message.from_user.id):
        await message.answer('Требуется авторизация.')
        return
    try:
        clients = get_clients() or []
    except requests.RequestException as e:
        await message.answer(f"Ошибка при получении клиентов: {e}")
        return
    except Exception as e:
        await message.answer(f"Неожиданная ошибка при получении клиентов: {e}")
        return

    if not clients:
        await message.answer("Нет клиентов")
        return

    # Попытка создать .xlsx через openpyxl, иначе вернуть .csv
    try:
        from openpyxl import Workbook
        import tempfile
        import os
        import asyncio

        wb = Workbook()
        ws = wb.active
        ws.title = "Clients"
        # Match the site header: ФИО, Пол, Возраст, Телефон, Email
        ws.append(["ФИО", "Пол", "Возраст", "Телефон", "Email"])
        for c in clients:
            ws.append([
                c.get("fio") or "",
                c.get("gender") or "",
                c.get("age") or "",
                c.get("phone") or "",
                c.get("email") or "",
            ])

        tf = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        tf.close()
        wb.save(tf.name)
        # send file via direct Telegram HTTP API in a thread to avoid aiogram InputFile issues
        def _post_file(chat_id, path, filename=None):
            token = os.getenv('BOT_TOKEN')
            if not token:
                return {'ok': False, 'error': 'BOT_TOKEN not set'}
            url = f'https://api.telegram.org/bot{token}/sendDocument'
            try:
                with open(path, 'rb') as f:
                    files = {'document': (filename or os.path.basename(path), f)}
                    resp = requests.post(url, data={'chat_id': chat_id}, files=files, timeout=30)
                return resp.json() if resp.ok else {'ok': False, 'error': resp.text}
            except Exception as e:
                return {'ok': False, 'error': str(e)}

        res = await asyncio.to_thread(_post_file, message.chat.id, tf.name, 'clients.xlsx')
        try:
            if not res.get('ok'):
                await message.answer(f"Ошибка отправки файла: {res.get('error')}")
        finally:
            try:
                os.remove(tf.name)
            except Exception:
                pass
        return
    except Exception:
        # Fall back to CSV if openpyxl is not available or save failed
        import tempfile, os
        tf = tempfile.NamedTemporaryFile(delete=False, suffix='.csv', mode='w', encoding='utf-8', newline='')
        writer = csv.writer(tf)
        # CSV header matching site
        writer.writerow(["ФИО", "Пол", "Возраст", "Телефон", "Email"])
        for c in clients:
            writer.writerow([
                c.get("fio") or "",
                c.get("gender") or "",
                c.get("age") or "",
                c.get("phone") or "",
                c.get("email") or "",
            ])
        tf.close()
        def _post_file(chat_id, path, filename=None):
            token = os.getenv('BOT_TOKEN')
            if not token:
                return {'ok': False, 'error': 'BOT_TOKEN not set'}
            url = f'https://api.telegram.org/bot{token}/sendDocument'
            try:
                with open(path, 'rb') as f:
                    files = {'document': (filename or os.path.basename(path), f)}
                    resp = requests.post(url, data={'chat_id': chat_id}, files=files, timeout=30)
                return resp.json() if resp.ok else {'ok': False, 'error': resp.text}
            except Exception as e:
                return {'ok': False, 'error': str(e)}

        res = await asyncio.to_thread(_post_file, message.chat.id, tf.name, 'clients.csv')
        try:
            if not res.get('ok'):
                await message.answer(f"Ошибка отправки файла: {res.get('error')}")
        finally:
            try:
                os.remove(tf.name)
            except Exception:
                pass


async def new_client(message: Message):
    await message.answer("Введите имя и телефон через запятую:")
    # дальше можно реализовать FSM


async def import_command(message: Message):
    """Start import flow: ask user to send a CSV or XLSX file."""
    if not _is_authorized(message.from_user.id):
        await message.answer('Требуется авторизация.')
        return
    _USER_STATE[message.from_user.id] = {'action': 'import'}
    await message.answer(
        "Отправьте файл (.csv или .xlsx) с колонками ФИО, Пол, Возраст, Телефон, Email.\n"
        "Бот обновит существующих клиентов по полю ФИО и попытается добавить новые записи.")


async def import_button_handler(message: Message):
    """Handler for the import button in the keyboard."""
    # reuse import flow
    await import_command(message)


async def _download_telegram_file(file_id: str) -> Tuple[str, str]:
    """Download a Telegram file by file_id. Returns (temp_path, filename)."""
    token = os.getenv('BOT_TOKEN')
    if not token:
        raise RuntimeError('BOT_TOKEN not set')
    info_url = f'https://api.telegram.org/bot{token}/getFile?file_id={file_id}'
    r = requests.get(info_url, timeout=20)
    r.raise_for_status()
    info = r.json()
    file_path = info.get('result', {}).get('file_path')
    if not file_path:
        raise RuntimeError('Не удалось получить file_path')
    file_url = f'https://api.telegram.org/file/bot{token}/{file_path}'
    r2 = requests.get(file_url, stream=True, timeout=60)
    r2.raise_for_status()
    # choose suffix from file_path
    filename = os.path.basename(file_path)
    tf = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(filename)[1])
    with open(tf.name, 'wb') as f:
        for chunk in r2.iter_content(8192):
            if chunk:
                f.write(chunk)
    return tf.name, filename


def _normalize_header(h: str) -> str:
    if not h:
        return h
    h2 = h.strip().lower()
    mapping = {
        'фио': 'fio', 'fio': 'fio', 'name': 'fio',
        'пол': 'gender', 'gender': 'gender',
        'возраст': 'age', 'age': 'age',
        'телефон': 'phone', 'phone': 'phone',
        'email': 'email', 'e-mail': 'email'
    }
    return mapping.get(h2, h2)


async def document_handler(message: Message):
    """Handle incoming document when user is in import state."""
    if not _is_authorized(message.from_user.id):
        await message.answer('Требуется авторизация.')
        return
    state = _USER_STATE.get(message.from_user.id)
    if not state or state.get('action') != 'import':
        await message.answer('Чтобы импортировать файл, сначала нажмите кнопку "📤 Импорт файла" или отправьте /import')
        return

    doc = getattr(message, 'document', None)
    if not doc:
        await message.answer('Не найден документ в сообщении')
        return

    await message.answer('Скачиваю файл...')
    try:
        tmp_path, filename = await asyncio.to_thread(_download_telegram_file, doc.file_id)
    except Exception as e:
        await message.answer(f'Ошибка при скачивании файла: {e}')
        _USER_STATE.pop(message.from_user.id, None)
        return

    updated = 0
    added = 0
    failed = 0
    try:
        ext = os.path.splitext(filename)[1].lower()
        rows = []
        if ext in ('.csv',):
            with open(tmp_path, 'r', encoding='utf-8-sig', newline='') as f:
                reader = csv.DictReader(f)
                headers = [ _normalize_header(h) for h in reader.fieldnames or [] ]
                for r in reader:
                    row = { headers[i]: v.strip() for i, v in enumerate(r.values()) if v is not None }
                    rows.append(row)
        else:
            # try openpyxl
            try:
                from openpyxl import load_workbook
            except Exception:
                await message.answer('Для обработки xlsx требуется пакет openpyxl')
                _USER_STATE.pop(message.from_user.id, None)
                return
            wb = load_workbook(tmp_path, read_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            headers_raw = next(it, None)
            headers = [ _normalize_header(h) for h in (headers_raw or []) ]
            for rr in it:
                row = { headers[i]: (rr[i] if rr and i < len(rr) else None) for i in range(len(headers)) }
                rows.append({k: (v if v is not None else '') for k, v in row.items()})

        # process rows
        for r in rows:
            fio = (r.get('fio') or '').strip()
            if not fio:
                failed += 1
                continue
            # build payload of non-empty fields except fio
            payload = {}
            for k in ('gender', 'age', 'phone', 'email'):
                if r.get(k) not in (None, ''):
                    payload[k] = r.get(k)
            try:
                # try update, if fails try add
                await asyncio.to_thread(update_person, fio, payload)
                updated += 1
            except Exception:
                try:
                    ph = r.get('phone') or ''
                    await asyncio.to_thread(add_client, fio, ph)
                    added += 1
                except Exception:
                    failed += 1
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
        _USER_STATE.pop(message.from_user.id, None)

    await message.answer(f'Импорт завершён. Обновлено: {updated}, добавлено: {added}, не обработано: {failed}')


async def _build_clients_keyboard(clients, page=0, page_size=8):
    kb = types.InlineKeyboardMarkup()
    start = page * page_size
    for c in clients[start:start + page_size]:
        fio = c.get('fio') or c.get('name') or ''
        btn = types.InlineKeyboardButton(text=fio, callback_data=f"view:{quote_plus(fio)}")
        kb.add(btn)
    # pagination
    if len(clients) > page_size:
        nav = []
        if page > 0:
            nav.append(types.InlineKeyboardButton('◀️', callback_data=f'page:{page-1}'))
        if (page+1)*page_size < len(clients):
            nav.append(types.InlineKeyboardButton('▶️', callback_data=f'page:{page+1}'))
        kb.row(*nav)
    return kb


async def list_clients_inline(message: Message):
    try:
        clients = get_clients() or []
    except Exception as e:
        await message.answer(f"Ошибка при получении клиентов: {e}")
        return
    if not clients:
        await message.answer('Нет клиентов')
        return
    kb = await _build_clients_keyboard(clients)
    await message.answer('Выберите клиента:', reply_markup=kb)


async def callback_handler(callback: types.CallbackQuery):
    data = callback.data or ''
    user_id = callback.from_user.id
    await callback.answer()
    if data.startswith('view:'):
        fio = unquote_plus(data.split(':', 1)[1])
        try:
            person = get_person(fio)
        except Exception as e:
            await callback.message.answer(f'Ошибка: {e}')
            return
        text = '\n'.join(f"{k}: {v}" for k, v in person.items())
        kb = types.InlineKeyboardMarkup(row_width=3)
        kb.add(
            types.InlineKeyboardButton('Редактировать', callback_data=f'edit:{quote_plus(fio)}'),
            types.InlineKeyboardButton('Заказы', callback_data=f'orders:{quote_plus(fio)}'),
            types.InlineKeyboardButton('Удалить', callback_data=f'delete:{quote_plus(fio)}')
        )
        await callback.message.answer(text, reply_markup=kb)
    elif data.startswith('orders:'):
        fio = unquote_plus(data.split(':', 1)[1])
        try:
            res = get_orders(fio)
            orders = res.get('orders') if isinstance(res, dict) else res
        except Exception as e:
            await callback.message.answer(f'Ошибка при получении заказов: {e}')
            return
        if not orders:
            await callback.message.answer('Заказов нет')
            return
        for o in orders:
            await callback.message.answer('\n'.join(f"{k}: {v}" for k, v in o.items()))
    elif data.startswith('edit:'):
        fio = unquote_plus(data.split(':', 1)[1])
        _USER_STATE[user_id] = {'action': 'edit', 'fio': fio}
        await callback.message.answer('Отправьте обновлённые поля в формате: поле: значение (например, phone: +7900...)\nМожно несколько в строках.')
    elif data.startswith('delete:'):
        fio = unquote_plus(data.split(':', 1)[1])
        try:
            delete_person(fio)
            await callback.message.answer('Клиент удалён')
        except Exception as e:
            await callback.message.answer(f'Ошибка при удалении: {e}')
    elif data.startswith('page:'):
        # rebuild page
        page = int(data.split(':', 1)[1])
        try:
            clients = get_clients() or []
        except Exception as e:
            await callback.message.answer(f"Ошибка: {e}")
            return
        kb = await _build_clients_keyboard(clients, page=page)
        await callback.message.edit_reply_markup(reply_markup=kb)


async def text_message_router(message: Message):
    """Route plain text messages depending on user in-memory state."""
    state = _USER_STATE.get(message.from_user.id)
    if not state:
        return
    if state.get('action') == 'edit':
        fio = state.get('fio')
        # parse lines like 'field: value'
        lines = [l.strip() for l in message.text.splitlines() if l.strip()]
        data = {}
        for ln in lines:
            if ':' in ln:
                k, v = ln.split(':', 1)
                data[k.strip()] = v.strip()
        if not data:
            await message.answer('Не удалось распознать поля. Отправьте в формате: phone: +7...')
            return
        try:
            update_person(fio, data)
            await message.answer('Данные обновлены')
        except Exception as e:
            await message.answer(f'Ошибка при обновлении: {e}')
        finally:
            _USER_STATE.pop(message.from_user.id, None)


async def login_command(message: Message):
    """Simple login: /login <user> <pass>"""
    text = (getattr(message, 'text', '') or '').strip()
    parts = text.split()
    if len(parts) < 3:
        await message.answer('Использование: /login <user> <pass>')
        return
    user = parts[1]
    pwd = parts[2]
    if user == _ADMIN_LOGIN and pwd == _ADMIN_PASS:
        _AUTH_USERS.add(message.from_user.id)
        await message.answer('Авторизация успешна. Вы можете использовать команды бота.', reply_markup=main_keyboard)
    else:
        await message.answer('Неверные учётные данные')


async def logout_command(message: Message):
    if message.from_user.id in _AUTH_USERS:
        _AUTH_USERS.discard(message.from_user.id)
        await message.answer('Вы вышли из системы')
    else:
        await message.answer('Вы не были авторизованы')
